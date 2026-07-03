import json
import os
import re
import sys
from pathlib import Path
from time import sleep

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

OUTPUT_DIR = "output"
DEFAULT_EXCEL_FILE = "persons.xlsx"
MAX_NAME_LENGTH = 50
INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')
NAZK_HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "uk-UA,uk;q=0.9,en;q=0.8",
    "Referer": "https://public.nazk.gov.ua/",
    "Origin": "https://public.nazk.gov.ua",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
}


def get_excel_path() -> str:
    args = [arg for arg in sys.argv[1:] if arg != "--check-only"]
    if args:
        return args[0]
    return os.environ.get("EXCEL_FILE", DEFAULT_EXCEL_FILE)


def sanitize_filename(name: str) -> str:
    safe_name = INVALID_FILENAME_CHARS.sub("_", name.strip())
    if len(safe_name) > MAX_NAME_LENGTH:
        safe_name = safe_name[:MAX_NAME_LENGTH].rstrip(" _")
    return safe_name


def build_output_path(edrpou: str, name: str) -> str:
    safe_name = sanitize_filename(name)
    filename = f"{edrpou}_{safe_name}.json"
    return os.path.join(OUTPUT_DIR, filename)


def normalize_edrpou(value) -> str | None:
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
        if not text:
            return None
        if "." in text and text.replace(".", "", 1).isdigit():
            text = str(int(float(text)))

    if text.isdigit():
        return text.zfill(8)

    return None


def create_excel_template(path: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ЄДРПОУ", "Назва"])
    sheet.append([39751280, "Приклад організації"])
    sheet.append([40381452, "НАЗК"])
    workbook.save(path)
    print(f"Створено шаблон {path}")


def read_persons_from_excel(path: str) -> list[tuple[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    persons = []

    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        edrpou = normalize_edrpou(row[0])
        if edrpou is None:
            continue

        name = str(row[1]).strip() if row[1] is not None else ""
        if not name:
            name = edrpou

        persons.append((edrpou, name))

    workbook.close()
    return persons


def check_nazk_access() -> bool:
    host = os.environ.get("NAZK_API_URL", "public-api.nazk.gov.ua")
    url = f"https://{host}/v2/documents/list"
    params = {"workPlaceEdrpou": "40381452", "page": 1}

    print("=== Перевірка доступу до NAZK API ===")
    print(f"Runner OS: {os.environ.get('RUNNER_OS', 'local')}")
    print(f"Runner name: {os.environ.get('RUNNER_NAME', 'local')}")
    print(f"GitHub Actions: {os.environ.get('GITHUB_ACTIONS', 'false')}")
    print(f"URL: {url}")

    session = requests.Session()
    session.headers.update(NAZK_HEADERS)
    response = session.get(url, params=params, timeout=30)

    print(f"HTTP status: {response.status_code}")

    if response.status_code == 200:
        print("РЕЗУЛЬТАТ: ДОСТУП Є — NAZK API відповідає")
        print("::notice::NAZK API доступний (HTTP 200)")
        return True

    if response.status_code == 403:
        print("РЕЗУЛЬТАТ: ЗАБЛОКОВАНО — NAZK API повернув 403 Forbidden")
        print("Ймовірна причина: IP GitHub datacenter заблоковано WAF/Cloudflare")
        print("Рішення: self-hosted runner на вашому ПК або локальний запуск")
        print("::error::NAZK API заблоковано (403 Forbidden)")
        return False

    print(f"РЕЗУЛЬТАТ: ПОМИЛКА — HTTP {response.status_code}")
    print(f"::error::NAZK API помилка (HTTP {response.status_code})")
    return False


def download_declarations(workplace_edrpou: str, output_file: str = None):
    base_url = os.environ.get("NAZK_API_URL")
    print(f"Downloading from {base_url}")
    base_url = f"https://{base_url}/v2/documents/list"

    page = 1
    all_items = []

    session = requests.Session()
    session.headers.update(NAZK_HEADERS)

    while True:
        params = {
            "workPlaceEdrpou": workplace_edrpou,
            "page": page
        }

        print(f"Отримання сторінки {page}...")

        response = session.get(base_url, params=params, timeout=30)
        response.raise_for_status()

        data = response.json()

        if isinstance(data, dict):
            if data.get("error") == 1310172:
                print("Всі сторінки отримано.")
                break
            if "error" in data:
                raise RuntimeError(f"NAZK API error: {data['error']}")

        if isinstance(data, list):
            all_items.extend(data)
        elif isinstance(data, dict):
            if "items" in data:
                all_items.extend(data["items"])
            elif "data" in data:
                all_items.extend(data["data"])
            elif "results" in data:
                all_items.extend(data["results"])
            else:
                all_items.append(data)

        page += 1
        sleep(0.2)

    if output_file is None:
        output_file = os.path.join(OUTPUT_DIR, f"nazk_{workplace_edrpou}.json")

    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(all_items, f, ensure_ascii=False, indent=4)

    print(f"Збережено {len(all_items)} записів у {output_file}")
    return len(all_items)


def process_excel(excel_path: str) -> int:
    persons = read_persons_from_excel(excel_path)
    print(f"Знайдено {len(persons)} організацій у {excel_path}")

    if not persons:
        print(f"У файлі {excel_path} не знайдено жодного валідного рядка.")
        return 0

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    total = len(persons)
    saved_count = 0

    for index, (edrpou, name) in enumerate(persons, start=1):
        output_file = build_output_path(edrpou, name)
        print(f"[{index}/{total}] Завантаження {edrpou} — {name}...")

        try:
            download_declarations(edrpou, output_file=output_file)
            saved_count += 1
        except Exception as error:
            print(f"Помилка для {edrpou} ({name}): {error}")

    print(f"Створено {saved_count} з {total} JSON-файлів у {OUTPUT_DIR}/")
    return saved_count


if __name__ == "__main__":
    check_only = "--check-only" in sys.argv

    if not check_nazk_access():
        sys.exit(1)

    if check_only:
        sys.exit(0)

    excel_path = get_excel_path()

    if not Path(excel_path).exists():
        create_excel_template(excel_path)

    saved_count = process_excel(excel_path)
    if saved_count == 0:
        sys.exit(1)
